"""读取 product 文件夹下的产品监控 Excel，构建 Product 模型。"""
from __future__ import annotations

import math
import os
import re
from typing import Optional

import openpyxl

from . import config
from .engine import normalize_code
from .models import Holding, Product


def parse_product_name_date(filename: str) -> tuple[str, str]:
    """从文件名解析产品名与日期。

    例：'稳进9号-实时监控20260610.xlsx' -> ('稳进9号', '20260610')
    规则：仅当末尾是【合法 YYYYMMDD 日期】(19/20 开头、月1-12、日1-31)时才剥离为日期，
          否则整体当产品名，避免把产品名里的合法数字(如 '组合12345678')误当日期。
    """
    base = os.path.splitext(os.path.basename(filename))[0]
    # 日期：末尾的 8 位合法日期（前面是字符串首或非数字字符，避免吃掉产品名数字）
    date = ""
    m = re.search(r"(?:^|[^0-9])((?:19|20)\d{2})(\d{2})(\d{2})\s*$", base)
    if m and 1 <= int(m.group(2)) <= 12 and 1 <= int(m.group(3)) <= 31:
        date = m.group(1) + m.group(2) + m.group(3)
        base = base[: m.start(1)]   # 用 start(1) 保留前置边界字符（如 '号'）
    # 去掉常见后缀词
    for sep in ["-实时监控", "_实时监控", "实时监控", "-监控", "监控", "-持仓", "持仓"]:
        if sep in base:
            base = base.split(sep)[0]
    # 去掉尾部分隔符
    name = base.strip(" -_·　")
    if not name:
        name = os.path.splitext(os.path.basename(filename))[0]
    return name, date


def _to_float(value, percent: bool = False) -> Optional[float]:
    """尽力把单元格值转为 float，'-'/None/空 视为 None。

    percent=True 时（仅「持仓权重」列）才把结尾 % 当百分比除以 100；其余列遇到结尾 %
    视为脏数据返回 None（避免价格/市值列里误格式化的 "12%" 被静默缩成 0.12 → 百倍超买）。
    同时容错券商导出的带币种/单位/量级文本（前导 ¥￥$/HK$、尾部 元/股/份、值级 万/亿）。
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return f if math.isfinite(f) else None   # 拒绝原生 nan/inf
    s = str(value).strip().replace(",", "").replace("，", "")
    if s == "" or s.lower() in ("-", "—", "/", "n/a", "nan", "inf", "-inf", "+inf", "infinity"):
        return None
    s = re.sub(r"^(?:hk\$|us\$|rmb|cny|hkd|usd|[¥￥$])\s*", "", s, flags=re.IGNORECASE)  # 前导币种符
    s = re.sub(r"\s*(?:元|股|份|手)$", "", s)                                              # 尾部单位字
    if s.endswith(("%", "％")):
        if not percent:
            return None                          # 价格/市值/数量列不接受百分比，按脏数据丢弃
        try:
            f = float(s[:-1]) / 100.0
        except ValueError:
            return None
        return f if math.isfinite(f) else None
    mult = 1.0                                    # 值级量级单位：'3.5万'→×1e4、'2亿'→×1e8
    m = re.search(r"(亿|万)$", s)
    if m:
        mult = 1e8 if m.group(1) == "亿" else 1e4
        s = s[:-1]
    try:
        f = float(s) * mult
    except ValueError:
        return None
    return f if math.isfinite(f) else None   # 拒绝 nan/inf，避免污染市值/价格/数量


def _norm_header(h) -> str:
    """归一化表头：去不间断空格/全角空格、去尾部括号单位，便于容错匹配。"""
    if h is None:
        return ""
    s = str(h).replace("\xa0", "").replace("　", "").strip()
    s = re.sub(r"[\(（][^\)）]*[\)）]\s*$", "", s)  # 去掉尾部 (CNY)/(元) 之类
    return s.strip()


def _clean_category(text: str) -> str:
    """清理分类文本：去掉不间断空格与计数括号，如 '   股票(136)' -> '股票'。"""
    if text is None:
        return ""
    t = str(text).replace("\xa0", "").strip()
    t = re.sub(r"[\(（]\d+[\)）]\s*$", "", t)
    return t.strip()


def _unit_scale(raw_header) -> float:
    """从表头单位后缀识别数量级倍率：万/万元→1e4，亿/亿元→1e8，否则 1。

    用于把「持仓市值（万元）」这类列的数值还原为原始单位（元），杜绝小 1 万倍单价 → 万倍超买。
    """
    if raw_header is None:
        return 1.0
    s = str(raw_header)
    m = re.search(r"[（(]\s*([^）)]*)\s*[)）]\s*$", s)
    unit = m.group(1) if m else ""
    if "亿" in unit:
        return 1e8
    if "万" in unit:
        return 1e4
    return 1.0


def _find_header_row(rows, max_probe: int = 6) -> int:
    """在前若干行里探测真正的表头行（匹配到最多必需列的那行）。

    应对首行是「标题/日期 banner」、真表头在第 2 行的布局——硬认 rows[0] 会把整只产品读空。
    """
    best_i, best_score = 0, -1
    for idx in range(min(max_probe, len(rows))):
        r = rows[idx]
        if r is None:
            continue
        norm = [_norm_header(c) for c in r]
        score = 0
        for aliases in config.HOLDING_COLUMN_ALIASES.values():
            na = [_norm_header(a) for a in aliases]
            if any(h and h in na for h in norm):
                score += 1
        if score > best_score:
            best_score, best_i = score, idx
    return best_i


def read_product_file(path: str) -> Product:
    """读取单个产品 Excel 文件。"""
    name, date = parse_product_name_date(path)
    product = Product(name=name, date=date, source_file=path)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if config.HOLDING_SHEET not in wb.sheetnames:
        # 兜底：用第一个 sheet
        ws = wb.worksheets[0]
        product.warnings.append(
            f"未找到名为「{config.HOLDING_SHEET}」的 sheet，已改用「{ws.title}」")
    else:
        ws = wb[config.HOLDING_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        product.warnings.append("持仓表为空")
        return product

    # 解析表头 -> 列索引（探测表头行；容忍不间断空格、尾部括号单位、列名别名）
    n_cols = max((len(r) for r in rows if r is not None), default=0)
    header_row = _find_header_row(rows)
    if header_row > 0:
        product.warnings.append(f"表头不在首行，已自动定位到第 {header_row + 1} 行（请核对）")
    hrow = rows[header_row] if header_row < len(rows) else ()
    norm_header = [_norm_header(hrow[i]) if i < len(hrow) else "" for i in range(n_cols)]
    col_idx: dict[str, int] = {}
    # 别名匹配带优先级：先在所有列里找【精确等于规范全名】(别名表首项，如 持仓数量/持仓市值/资产名称)，
    # 找不到再退到次级别名取最左列——否则短别名「持仓」会抢在「持仓数量」前绑错 quantity。
    for field_name, aliases in config.HOLDING_COLUMN_ALIASES.items():
        norm_aliases = [_norm_header(a) for a in aliases]
        canonical = norm_aliases[0]
        exact = next((i for i, h in enumerate(norm_header) if h and h == canonical), None)
        if exact is not None:
            col_idx[field_name] = exact
            continue
        for i, h in enumerate(norm_header):
            if h and h in norm_aliases:
                col_idx[field_name] = i
                break
    # 每列的数量级倍率（万元/亿元等），用于把市值/价格/数量还原到原始单位
    col_scale = {f: _unit_scale(hrow[i]) if i < len(hrow) else 1.0 for f, i in col_idx.items()}
    # 必需列校验
    for required in ("code", "name", "price", "quantity", "market_value"):
        if required not in col_idx:
            product.warnings.append(f"缺少必要列：{required}（表头未能匹配，可能列名异常）")
    has_cat_col = "category" in col_idx
    cat_col = col_idx.get("category", 0)
    if not has_cat_col:
        product.warnings.append("未匹配到「分类」列，已按代码列里的大类关键字识别汇总行")

    def cell(r, field):
        """安全取值：列缺失或行长不足均返回 None（兼容 read_only 的参差行）。"""
        i = col_idx.get(field)
        if i is None or i >= len(r):
            return None
        return r[i]

    def fnum(r, field):
        """读数值并按列单位倍率还原（市值/价格/数量）。"""
        v = _to_float(cell(r, field))
        return v * col_scale.get(field, 1.0) if v is not None else None

    n_guessed = 0
    current_category = ""
    for r in rows[header_row + 1:]:
        if r is None:
            continue
        code_raw = cell(r, "code")
        a = r[cat_col] if (has_cat_col and cat_col < len(r)) else None  # 分类列
        cat_text = _clean_category(a) if a is not None else ""
        code_text = _clean_category(code_raw) if code_raw is not None else ""

        # 大类汇总行：有分类列时＝分类列有值且代码列为空；
        #            无分类列时＝代码列【整词恰为】大类关键字（全部/股票/债券…）。
        # 必须整词相等，不能用子串：否则代码列里混入的证券名（如「国债逆回购991」「南方…ETF联接基金」
        # 含 回购/基金 子串）会被误当汇总行而把真实持仓静默丢弃。
        summary_label = ""
        if has_cat_col:
            if cat_text and (code_raw is None or str(code_raw).strip() == ""):
                summary_label = cat_text
        elif code_text in config.CATEGORY_KEYWORDS:
            summary_label = code_text
        if summary_label:
            current_category = summary_label
            mv = fnum(r, "market_value")
            if config.TOTAL_KEYWORD in summary_label and mv is not None:
                product.total_assets = mv
            if config.STOCK_KEYWORD in summary_label and config.TOTAL_KEYWORD not in summary_label and mv is not None:
                product.stock_total_mv = mv
            continue

        # 普通持仓行
        if code_raw is None or str(code_raw).strip() == "":
            continue
        code = str(code_raw).strip()
        norm, guessed = normalize_code(code)   # 补全市场后缀：600000→600000.SH、1→000001.SZ
        if norm:
            code = norm
            if guessed:
                n_guessed += 1
        name_raw = cell(r, "name")
        name_v = str(name_raw).strip() if name_raw is not None else ""
        price = fnum(r, "price")
        qty = fnum(r, "quantity")
        mv = fnum(r, "market_value")
        weight = _to_float(cell(r, "weight"), percent=True)

        product.holdings.append(
            Holding(
                code=code,
                name=name_v,
                category=current_category,
                price=price if price is not None else 0.0,
                quantity=qty if qty is not None else 0.0,
                market_value=mv if mv is not None else 0.0,
                weight=weight if weight is not None else 0.0,
            )
        )

    if n_guessed:
        product.warnings.append(f"{n_guessed} 个代码缺市场后缀，已自动推断补全（如 600000→600000.SH，请核对）")

    # 同一代码多行持仓：合并数量/市值（否则 by_code 仅留最后一行 → 漏卖、持仓比例失真）
    merged_h: dict[str, Holding] = {}
    order: list[str] = []
    has_dup = False
    for h in product.holdings:
        if h.code in merged_h:
            has_dup = True
            m = merged_h[h.code]
            m.quantity += h.quantity
            m.market_value += h.market_value
            m.weight += h.weight
            if not m.price and h.price:
                m.price = h.price
            if not m.name and h.name:
                m.name = h.name
            if not m.category and h.category:
                m.category = h.category
        else:
            merged_h[h.code] = h
            order.append(h.code)
    if has_dup:
        product.holdings = [merged_h[c] for c in order]
        product.warnings.append("检测到同一代码多行持仓，已合并其数量与市值（请核对）")

    product.rebuild_index()

    # 兜底：若没解析到总资产/股票市值，用持仓市值之和（与汇总行可能略有出入，给出告警）
    if not product.total_assets:
        product.total_assets = sum(h.market_value for h in product.holdings)
        if product.total_assets:
            product.warnings.append(
                "未读到「全部」汇总行的总资产，已用各持仓市值之和估算（按比例调仓时请人工复核）")
    if not product.stock_total_mv:
        product.stock_total_mv = sum(h.market_value for h in product.stocks())

    return product


def read_product_folder(folder: str) -> list[Product]:
    """读取文件夹下所有 .xlsx 产品文件（忽略 ~$ 临时文件）。"""
    products: list[Product] = []
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"文件夹不存在：{folder}")
    files = sorted(
        f for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xlsm"))
        and not f.startswith("~$")
        and not f.startswith("交易指令")        # 跳过本工具导出的下单指令，避免被误当成产品
        and not f.startswith("调仓输入")
    )
    for f in files:
        path = os.path.join(folder, f)
        try:
            products.append(read_product_file(path))
        except Exception as exc:  # noqa: BLE001 单个文件失败不影响其他
            p = Product(name=os.path.splitext(f)[0], source_file=path)
            p.warnings.append(f"读取失败：{exc}")
            products.append(p)
    # 同名产品去重：否则按名建键时后者覆盖前者，前一只的持仓会凭空消失
    seen: dict[str, int] = {}
    for p in products:
        if p.name in seen:
            seen[p.name] += 1
            stem = os.path.splitext(os.path.basename(p.source_file))[0]
            p.name = f"{p.name}（{stem}）" if stem else f"{p.name}#{seen[p.name]}"
            p.warnings.append("检测到同名产品，已用文件名区分，请核对是否重复")
        else:
            seen[p.name] = 1
    return products
