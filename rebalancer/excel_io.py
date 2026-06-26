"""Excel 写出：券商下单指令（调仓输入在界面上直接录入，不经 Excel）。"""
from __future__ import annotations

import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config
from .models import TradeOrder


# ---------------------------------------------------------------------------
# 写出券商下单指令
# ---------------------------------------------------------------------------
_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_THIN = Side(style="thin", color="808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")


def _write_product_sheet(ws, product_name: str, orders: list[TradeOrder]) -> None:
    # 表头：A 列放产品名，其后为交易方向/证券代码/证券名称/买入数量/卖出数量
    headers = [product_name] + config.OUTPUT_HEADERS
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER

    # 卖出在前、买入在后（引擎已排序，这里再保证一次）
    dir_rank = {config.DIR_SELL: 0, config.DIR_BUY: 1}
    ordered = sorted(orders, key=lambda o: (dir_rank.get(o.direction, 9), o.code))

    for i, o in enumerate(ordered, start=1):
        r = i + 1
        values = [
            i,                                          # 序号
            o.direction,                                # 交易方向
            o.code,                                      # 证券代码
            o.name,                                      # 证券名称
            int(o.buy_qty) if o.buy_qty else None,       # 买入数量
            int(o.sell_qty) if o.sell_qty else None,     # 卖出数量
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = _CENTER
            cell.border = _BORDER
            if o.direction == config.DIR_SELL:
                cell.fill = _YELLOW

    # 列宽
    widths = [10, 10, 14, 16, 12, 12]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def safe_filename(name: str) -> str:
    """统一的文件名净化（去掉 Windows/通用非法字符），供导出命名与覆盖预判共用。"""
    return re.sub(r'[\\/:*?"<>|\[\]]', "_", name)


def _safe_sheet_title(name: str, used: set) -> str:
    title = re.sub(r"[\[\]\*\?:/\\]", "_", name)[:31] or "Sheet"
    base = title
    i = 1
    while title in used:
        i += 1
        suffix = f"_{i}"
        title = base[: 31 - len(suffix)] + suffix
    used.add(title)
    return title


def write_orders(orders: list[TradeOrder], path: str) -> list[str]:
    """把交易指令写到一个工作簿，每个产品一个 sheet。返回涉及的产品名列表。"""
    products: list[str] = []
    for o in orders:
        if o.product not in products:
            products.append(o.product)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_titles: set = set()
    if not products:
        ws = wb.create_sheet("无交易")
        ws.cell(row=1, column=1, value="无需调仓（没有产生任何交易指令）")
    for p in products:
        ws = wb.create_sheet(_safe_sheet_title(p, used_titles))
        _write_product_sheet(ws, p, [o for o in orders if o.product == p])
    wb.save(path)
    return products


def write_orders_per_product(orders: list[TradeOrder], folder: str, date: str = "") -> list[str]:
    """每个产品单独导出一个文件，便于分别发券商。返回生成的文件路径列表。"""
    os.makedirs(folder, exist_ok=True)
    products: list[str] = []
    for o in orders:
        if o.product not in products:
            products.append(o.product)
    paths: list[str] = []
    used_names: set = set()
    suffix = f"_{date}" if date else ""
    for p in products:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _safe_sheet_title(p, set())
        _write_product_sheet(ws, p, [o for o in orders if o.product == p])
        safe_name = safe_filename(p)
        base = f"交易指令_{safe_name}{suffix}"
        # 文件名去重：不同产品名净化后可能撞名，加序号避免互相覆盖丢单
        fname = base
        k = 1
        while fname in used_names:
            k += 1
            fname = f"{base}_{k}"
        used_names.add(fname)
        path = os.path.join(folder, fname + ".xlsx")
        wb.save(path)
        paths.append(path)
    if len(set(paths)) != len(products):   # 去重逻辑已保证唯一；显式兜底，-O 下仍生效（不用 assert）
        raise RuntimeError("导出文件名去重失败：存在重名文件，可能互相覆盖")
    return paths
