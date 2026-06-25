"""Excel 输出测试（券商下单指令）。调仓输入在界面直接录入，不经 Excel。"""
import os

import openpyxl

from rebalancer import (DirectionGroup, StockEntry, build_orders, write_orders,
                        write_orders_per_product)
from rebalancer import config
from rebalancer.models import TradeOrder


def test_write_orders_format(tmp_path, synth_product, synth_merged):
    gb = DirectionGroup("测试1号", "买入", 100000, "等金额", [StockEntry("600001.SH", price=20.0)])
    gs = DirectionGroup("测试1号", "卖出", 30000, "等金额", [StockEntry("600000.SH", price=10.0)])
    res = build_orders([gb, gs], [synth_product], synth_merged)
    path = str(tmp_path / "下单.xlsx")
    write_orders(res.orders, path)
    wb = openpyxl.load_workbook(path)
    ws = wb["测试1号"]
    assert ws["A1"].value == "测试1号"
    assert [ws.cell(1, c).value for c in range(2, 7)] == config.OUTPUT_HEADERS
    # 第一条卖出且标黄
    assert ws.cell(2, 2).value == config.DIR_SELL
    assert ws.cell(2, 2).fill.start_color.rgb in ("00FFFF00", "FFFFFF00")


def test_export_filename_dedup(tmp_path):
    orders = [
        TradeOrder(product="A/B", direction=config.DIR_BUY, code="600000.SH", name="甲", buy_qty=100),
        TradeOrder(product="A_B", direction=config.DIR_BUY, code="600001.SH", name="乙", buy_qty=100),
    ]
    paths = write_orders_per_product(orders, str(tmp_path))
    assert len(paths) == 2 and len(set(paths)) == 2
    for p in paths:
        assert os.path.exists(p)
